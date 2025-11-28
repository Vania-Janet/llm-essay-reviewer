#!/usr/bin/env python3
"""
Script para generar un Excel profesional de ensayos evaluados directamente desde la base de datos.
"""
from pathlib import Path
import sys
from datetime import datetime

# Agregar el directorio raíz al path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("WARN: openpyxl no esta instalado. Instalando...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True

# Importar modelos de base de datos
from flask import Flask
from app.database.connection import db
from app.database.models import Ensayo
from app.config import Config

# Crear app Flask
app = Flask(__name__)
app.config.from_object(Config)
db.init_app(app)

def obtener_ensayos():
    """Obtiene todos los ensayos de la base de datos."""
    with app.app_context():
        ensayos = Ensayo.query.order_by(Ensayo.puntuacion_total.desc()).all()
        return ensayos

def crear_excel_profesional(output_path: str = None):
    """
    Crea un Excel profesional con formato mejorado desde la base de datos.
    
    Args:
        output_path: Ruta de salida para el Excel (opcional)
    """
    print(f"\n{'='*70}")
    print("GENERANDO EXCEL PROFESIONAL")
    print(f"{'='*70}\n")
    
    # Obtener ensayos de la base de datos
    print(f"Cargando ensayos de la base de datos...")
    ensayos = obtener_ensayos()
    print(f"{len(ensayos)} ensayos cargados\n")
    
    if not ensayos:
        print("ERROR: No hay ensayos en la base de datos")
        return None
    
    # Determinar ruta de salida
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = Path.cwd() / f"ensayos_evaluados_{timestamp}_profesional.xlsx"
    
    # Crear workbook
    print(f"Guardando en: {output_path}")
    wb = Workbook()
    ws = wb.active
    ws.title = 'Evaluaciones'
    
    # Encabezados
    headers = [
        'Ranking', 'Puntuación Total', 'Nombre de Archivo', 'Autor',
        'Calidad Técnica', 'Creatividad', 'Vinculación Temática',
        'Bienestar Colectivo', 'Uso Responsable IA', 'Potencial Impacto',
        'Tiene Anexo', 'Fecha Evaluación', 'Longitud (palabras)', 'Comentario General'
    ]
    ws.append(headers)
    
    # Agregar datos
    for idx, ensayo in enumerate(ensayos, start=1):
        # Los campos ya están en el modelo como JSON
        calidad_tecnica = ensayo.calidad_tecnica or {}
        creatividad = ensayo.creatividad or {}
        vinculacion_tematica = ensayo.vinculacion_tematica or {}
        bienestar_colectivo = ensayo.bienestar_colectivo or {}
        uso_responsable_ia = ensayo.uso_responsable_ia or {}
        potencial_impacto = ensayo.potencial_impacto or {}
        
        # Extraer autor del nombre del archivo
        autor = ensayo.autor or 'N/A'
        
        row_data = [
            idx,  # Ranking
            round(ensayo.puntuacion_total, 2),
            ensayo.nombre_archivo_original or ensayo.nombre_archivo,
            autor,
            calidad_tecnica.get('calificacion', 'N/A'),
            creatividad.get('calificacion', 'N/A'),
            vinculacion_tematica.get('calificacion', 'N/A'),
            bienestar_colectivo.get('calificacion', 'N/A'),
            uso_responsable_ia.get('calificacion', 'N/A'),
            potencial_impacto.get('calificacion', 'N/A'),
            'Sí' if ensayo.tiene_anexo else 'No',
            ensayo.fecha_evaluacion.strftime('%Y-%m-%d %H:%M') if ensayo.fecha_evaluacion else 'N/A',
            ensayo.num_palabras or (len(ensayo.texto_completo.split()) if ensayo.texto_completo else 0),
            ensayo.comentario_general or 'N/A'
        ]
        ws.append(row_data)
    
    print("Aplicando formato profesional...")
    
    # Colores corporativos (azul profesional y gris)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    alt_row_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Colores para puntuaciones
    excellent_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Verde claro
    good_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Amarillo claro
    poor_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Rojo claro
    
    # Estilos de borde
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # 1. Formatear encabezado
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=11, name='Calibri')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    # 2. Ajustar anchos de columna
    column_widths = {
        'A': 10,  # Ranking
        'B': 15,  # Puntuación Total
        'C': 50,  # Nombre de Archivo
        'D': 30,  # Autor
        'E': 15,  # Calidad Técnica
        'F': 15,  # Creatividad
        'G': 18,  # Vinculación Temática
        'H': 18,  # Bienestar Colectivo
        'I': 18,  # Uso Responsable IA
        'J': 15,  # Potencial Impacto
        'K': 12,  # Tiene Anexo
        'L': 18,  # Fecha Evaluación
        'M': 15,  # Longitud (palabras)
        'N': 60,  # Comentario General
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 3. Formatear filas de datos
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        # Alternar color de fila
        if row_idx % 2 == 0:
            for cell in row:
                if not cell.fill.start_color.rgb or cell.fill.start_color.rgb == '00000000':
                    cell.fill = alt_row_fill
        
        # Aplicar bordes y alineación
        for cell in row:
            cell.border = thin_border
            cell.font = Font(name='Calibri', size=10)
            
            # Alineación específica por columna
            if cell.column_letter in ['A', 'B', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'M']:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif cell.column_letter in ['C', 'D', 'N']:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            elif cell.column_letter == 'L':
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 4. Aplicar formato condicional a puntuación total (columna B)
        puntuacion_cell = row[1]  # Columna B
        if puntuacion_cell.value and isinstance(puntuacion_cell.value, (int, float)):
            puntuacion = float(puntuacion_cell.value)
            if puntuacion >= 4.5:
                puntuacion_cell.fill = excellent_fill
                puntuacion_cell.font = Font(bold=True, color="006100", size=11, name='Calibri')
            elif puntuacion >= 3.5:
                puntuacion_cell.fill = good_fill
                puntuacion_cell.font = Font(bold=True, color="9C6500", size=11, name='Calibri')
            else:
                puntuacion_cell.fill = poor_fill
                puntuacion_cell.font = Font(bold=True, color="9C0006", size=11, name='Calibri')
        
        # 5. Formatear ranking (columna A)
        ranking_cell = row[0]
        if ranking_cell.value:
            ranking_cell.font = Font(bold=True, size=11, name='Calibri')
        
        # 6. Formatear "Tiene Anexo" (columna K)
        anexo_cell = row[10]  # Columna K
        if anexo_cell.value == 'Sí':
            anexo_cell.font = Font(color="006100", bold=True, name='Calibri')
        elif anexo_cell.value == 'No':
            anexo_cell.font = Font(color="9C0006", name='Calibri')
    
    # 7. Ajustar altura de filas
    ws.row_dimensions[1].height = 40  # Encabezado más alto
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 50  # Filas de datos
    
    # 8. Agregar filtros automáticos
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    
    # 9. Congelar primera fila y primera columna
    ws.freeze_panes = "B2"
    
    # Guardar
    wb.save(output_path)
    print(f"Excel profesional guardado exitosamente\n")
    
    # Calcular estadísticas
    con_anexo = sum(1 for e in ensayos if e.tiene_anexo)
    sin_anexo = len(ensayos) - con_anexo
    puntuaciones = [e.puntuacion_total for e in ensayos]
    promedio = sum(puntuaciones) / len(puntuaciones) if puntuaciones else 0
    
    # Estadísticas
    print(f"{'='*70}")
    print("ESTADISTICAS DEL ARCHIVO")
    print(f"{'='*70}")
    print(f"Total de ensayos: {len(ensayos)}")
    print(f"Con anexo: {con_anexo}")
    print(f"Sin anexo: {sin_anexo}")
    print(f"Puntuacion promedio: {promedio:.2f}")
    print(f"Puntuacion maxima: {max(puntuaciones):.2f}")
    print(f"Puntuacion minima: {min(puntuaciones):.2f}")
    print(f"\n{'='*70}")
    print("ARCHIVO EXCEL PROFESIONAL GENERADO")
    print(f"{'='*70}\n")
    
    return output_path

def main():
    """Función principal."""
    print(f"Generando Excel desde base de datos...")
    
    # Generar Excel
    output_path = crear_excel_profesional()
    
    if output_path:
        print(f"Listo! Abre el archivo: {output_path}")
        return 0
    else:
        print("ERROR: No se pudo generar el archivo Excel")
        return 1

if __name__ == '__main__':
    sys.exit(main())
